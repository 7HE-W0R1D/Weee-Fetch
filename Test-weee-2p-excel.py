import requests
import json
import time
import io
import os
try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage
    from tqdm import tqdm
except ImportError:
    print("缺少必要的库，请在终端运行: pip install openpyxl pillow tqdm")
    exit()

# 引入公共请求头工具
import weee_headers

def fetch_all_reviews(product_id, max_pages=2, lang='zh'):
    base_url = "https://api.sayweee.net/ec/social/review"
    all_reviews = []
    page = 1
    limit = 20
    
    with tqdm(total=max_pages, desc="Fetching review pages", unit="page") as pbar:
        while page <= max_pages:
            params = {
                "product_id": product_id,
                "sort": "relevance",
                "page": page,
                "limit": limit
            }
            
            try:
                response = requests.get(base_url, params=params, headers=weee_headers.get_headers(product_id, lang))
                data = response.json()
                
                object_data = data.get('object') or {}
                reviews = object_data.get('list', [])
                
                if not reviews:
                    break
                    
                all_reviews.extend(reviews)
                
                page += 1
                pbar.update(1)
                time.sleep(1) 
                
            except Exception as e:
                pbar.write(f"Error at page {page}: {e}")
                break
                
    return all_reviews

# 开始抓取
import argparse

parser = argparse.ArgumentParser(description='Fetch Weee reviews (2p Excel)')
parser.add_argument('--product-id', default='27605', help='Product ID to fetch')
parser.add_argument('--lang', default='zh', choices=['zh', 'en'], help='Language (zh or en)')
args = parser.parse_args()
product_id = args.product_id
lang = args.lang

data = fetch_all_reviews(product_id, lang=lang)

current_time = time.strftime("%Y%m%d_%H%M%S")
num_reviews = len(data)
base_name = f"{product_id}_{current_time}_{num_reviews}"
base_dir = os.path.join("fetched-data", base_name)

if not os.path.exists(base_dir):
    os.makedirs(base_dir)

json_filename = os.path.join(base_dir, f"{base_name}.json")
excel_filename = os.path.join(base_dir, f"{base_name}.xlsx")

# 保存为JSON文件
with open(json_filename, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

# 保存为Excel文件
if data:
    print("开始生成并排版Excel，并在其中嵌入图片，请稍候...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"

    # 提取所有的列名
    fieldnames = []
    for review in data:
        for key in review:
            if key not in fieldnames:
                fieldnames.append(key)
    
    # 找到最多有多少张图
    max_pics = 0
    for review in data:
        pics = review.get('pictures')
        if isinstance(pics, list) and len(pics) > max_pics:
            max_pics = len(pics)

    # 移除原本的pictures字段文本列，改成动态图片列
    if 'pictures' in fieldnames:
        fieldnames.remove('pictures')
    
    headers = fieldnames + [f"picture_{i+1}" for i in range(max_pics)]
    ws.append(headers)

    # 调整图片列的宽度 (宽度36 大约像素250左右)
    for i in range(max_pics):
        col_letter = get_column_letter(len(fieldnames) + 1 + i)
        ws.column_dimensions[col_letter].width = 36

    for row_idx, review in enumerate(tqdm(data, desc="Processing rows and images"), start=2):
        # 调整所有数据行的行高 (行高190 大约像素253)
        ws.row_dimensions[row_idx].height = 190
        
        # 写入普通文本
        for col_idx, key in enumerate(fieldnames, start=1):
            val = review.get(key)
            if isinstance(val, (dict, list)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx).value = val
            
        # 写入并下载图片
        pictures = review.get('pictures')
        if isinstance(pictures, list):
            for pic_idx, img_url in enumerate(pictures):
                col_idx = len(fieldnames) + 1 + pic_idx
                if img_url.startswith('//'):
                    img_url = 'https:' + img_url
                    
                try:
                    img_response = requests.get(img_url, timeout=5)
                    if img_response.status_code == 200:
                        image_stream = io.BytesIO(img_response.content)
                        
                        pil_img = PILImage.open(image_stream)
                        # 获取原图尺寸，在保证清晰度的情况下只在 Excel 视图层面调整尺寸
                        width, height = pil_img.size
                        ratio = min(500/width, 500/height) if width > 0 and height > 0 else 1
                        
                        img_byte_arr = io.BytesIO()
                        # 防止RGBA透明通道保存为JPEG报错，统一转存PNG
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        
                        xl_img = OpenpyxlImage(img_byte_arr)
                        # 调整在 Excel 中的显示大小，而不是压缩图片像素（保持原图高清晰度）
                        xl_img.width = int(width * ratio)
                        xl_img.height = int(height * ratio)
                        # 设置锚点为对应单元格
                        xl_img.anchor = ws.cell(row=row_idx, column=col_idx).coordinate
                        ws.add_image(xl_img)
                    else:
                        ws.cell(row=row_idx, column=col_idx).value = "【图片失效】"
                except Exception as e:
                    print(f"Failed to load image for row {row_idx}: {e}")
                    # 如果下载或嵌入失败，将URL写在单元格里
                    ws.cell(row=row_idx, column=col_idx).value = img_url

    wb.save(excel_filename)
    print(f"全部完成! 文件已保存为 {json_filename} 和 {excel_filename}")
