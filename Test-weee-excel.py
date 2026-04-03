import requests
import json
import time
import io
import os
try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage
    from tqdm import tqdm
except ImportError:
    print("缺少必要的库，请在终端运行: pip install openpyxl pillow tqdm")
    exit()

# 引入公共请求头工具
import weee_headers

def fetch_all_reviews(product_id, max_reviews=None, lang='zh'):
    base_url = "https://api.sayweee.net/ec/social/review"
    all_reviews = []
    page = 1
    limit = 20
    
    with tqdm(desc="Fetching reviews", unit="review") as pbar:
        while True:
            params = {
                "product_id": product_id,
                "sort": "relevance",
                "page": page,
                "limit": limit
            }
            
            try:
                response = requests.get(base_url, params=params, headers=weee_headers.get_headers(product_id, lang))
                data = response.json()
                
                object_data = data.get('object') or {}
                reviews = object_data.get('list', [])
                
                if page == 1:
                    total_from_api = object_data.get('total')
                    if total_from_api:
                        pbar.total = min(total_from_api, max_reviews) if max_reviews else total_from_api
                        pbar.refresh()
                        
                if not reviews:
                    break
                    
                added_reviews = min(len(reviews), max_reviews - len(all_reviews) if max_reviews else len(reviews))
                all_reviews.extend(reviews[:added_reviews])
                pbar.update(added_reviews)
                
                if max_reviews is not None and len(all_reviews) >= max_reviews:
                    break
                    
                page += 1
                time.sleep(1) 
                
            except Exception as e:
                pbar.write(f"Error at page {page}: {e}")
                break
                
    if max_reviews is not None and len(all_reviews) >= max_reviews:
        print(f"Successfully fetched {len(all_reviews)} reviews (reached the requested limit).")
        
    return all_reviews

# 开始抓取
import argparse
import re

parser = argparse.ArgumentParser(description='Fetch Weee reviews to Excel')
parser.add_argument('--product-id', help='Product ID or full URL')
parser.add_argument('--lang', help='Language (cn for Chinese, en for English)')
args = parser.parse_args()

# Determine product_id and product_name
product_name = ""
if args.product_id:
    if args.product_id.startswith('http'):
        parts = args.product_id.split('?')[0].strip('/').split('/')
        product_id = parts[-1]
        product_name = parts[-2] if len(parts) >= 2 else ''
    else:
        product_id = args.product_id
        product_name = ''
else:
    url_input = input("请输入商品链接或直接输入商品ID (直接回车默认使用 27605): ").strip()
    if url_input:
        if url_input.startswith('http'):
            parts = url_input.split('?')[0].strip('/').split('/')
            product_id = parts[-1]
            product_name = parts[-2] if len(parts) >= 2 else ""
        else:
            product_id = url_input
    else:
        product_id = "27605"

# Determine language
if args.lang:
    lang = 'zh' if args.lang.lower() == 'cn' else 'en'
else:
    lang_input = input("请选择语言 (cn 为中文, en 为英文, 默认 cn): ").strip().lower()
    if lang_input == 'en':
        lang = 'en'
    else:
        lang = 'zh'
# Optional: you could still ask for product_name as before if needed

import re
if product_name:
    # 清洗掉文件名中不合法字符
    product_name = re.sub(r'[\\/:*?"<>| ]', '-', product_name)

try:
    user_input = input("请输入要抓取的最大评论数 (直接回车则默认抓取全部评论): ").strip()
    max_reviews = int(user_input) if user_input else None
except ValueError:
    print("输入无效，将默认抓取全部评论。")
    max_reviews = None

data = fetch_all_reviews(product_id, max_reviews=max_reviews, lang=lang)

current_time = time.strftime("%Y%m%d_%H%M%S")
num_reviews = len(data)
base_name = f"{product_name}_{product_id}_{current_time}_{num_reviews}" if product_name else f"{product_id}_{current_time}_{num_reviews}"
base_dir = os.path.join("fetched-data", base_name)

if not os.path.exists(base_dir):
    os.makedirs(base_dir)

json_filename = os.path.join(base_dir, f"{base_name}.json")
excel_filename = os.path.join(base_dir, f"{base_name}.xlsx")

# 保存为JSON文件
with open(json_filename, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

# 保存为Excel文件
if data:
    print("开始生成并排版Excel，并在其中嵌入图片，请稍候...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"

    # 提取所有的列名
    fieldnames = []
    for review in data:
        for key in review:
            if key not in fieldnames:
                fieldnames.append(key)
    
    # 找到最多有多少张图
    max_pics = 0
    for review in data:
        pics = review.get('pictures')
        if isinstance(pics, list) and len(pics) > max_pics:
            max_pics = len(pics)

    # 移除原本的pictures字段文本列，改成动态图片列
    if 'pictures' in fieldnames:
        fieldnames.remove('pictures')
    
    headers = fieldnames + [f"picture_{i+1}" for i in range(max_pics)]
    ws.append(headers)

    # 调整图片列的宽度 (宽度36 大约像素250左右)
    for i in range(max_pics):
        col_letter = get_column_letter(len(fieldnames) + 1 + i)
        ws.column_dimensions[col_letter].width = 36

    for row_idx, review in enumerate(tqdm(data, desc="Processing rows and images"), start=2):
        # 调整所有数据行的行高 (行高190 大约像素253)
        ws.row_dimensions[row_idx].height = 190
        
        # 写入普通文本
        for col_idx, key in enumerate(fieldnames, start=1):
            val = review.get(key)
            if isinstance(val, (dict, list)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx).value = val
            
        # 写入并下载图片
        pictures = review.get('pictures')
        if isinstance(pictures, list):
            for pic_idx, img_url in enumerate(pictures):
                col_idx = len(fieldnames) + 1 + pic_idx
                if img_url.startswith('//'):
                    img_url = 'https:' + img_url
                    
                try:
                    img_response = requests.get(img_url, timeout=5)
                    if img_response.status_code == 200:
                        image_stream = io.BytesIO(img_response.content)
                        
                        pil_img = PILImage.open(image_stream)
                        # 获取原图尺寸，在保证清晰度的情况下只在 Excel 视图层面调整尺寸
                        width, height = pil_img.size
                        ratio = min(500/width, 500/height) if width > 0 and height > 0 else 1
                        
                        img_byte_arr = io.BytesIO()
                        # 防止RGBA透明通道保存为JPEG报错，统一转存PNG
                        pil_img.save(img_byte_arr, format='PNG')
                        img_byte_arr.seek(0)
                        
                        xl_img = OpenpyxlImage(img_byte_arr)
                        # 调整在 Excel 中的显示大小，而不是压缩图片像素（保持原图高清晰度）
                        xl_img.width = int(width * ratio)
                        xl_img.height = int(height * ratio)
                        # 设置锚点为对应单元格
                        xl_img.anchor = ws.cell(row=row_idx, column=col_idx).coordinate
                        ws.add_image(xl_img)
                    else:
                        ws.cell(row=row_idx, column=col_idx).value = "【图片失效】"
                except Exception as e:
                    print(f"Failed to load image for row {row_idx}: {e}")
                    # 如果下载或嵌入失败，将URL写在单元格里
                    ws.cell(row=row_idx, column=col_idx).value = img_url

    wb.save(excel_filename)
    print(f"全部完成! 文件已保存为 {json_filename} 和 {excel_filename}")
